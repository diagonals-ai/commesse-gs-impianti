import streamlit as st
import pdfplumber
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io
import os
import pandas as pd
from pathlib import Path

st.set_page_config(
    page_title="GS Impianti – Generatore Commesse",
    page_icon="⚡",
    layout="wide",
)

# ── CSS ─────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
    .main-title { font-size:2rem; font-weight:700; color:#1F3864; margin-bottom:0; }
    .sub-title  { font-size:1rem; color:#666; margin-bottom:1rem; }
    .step-box   { background:#f0f4ff; border-left:4px solid #1F3864;
                  padding:1rem 1.2rem; border-radius:6px; margin-bottom:1rem; }
    .step-num   { font-weight:700; color:#1F3864; font-size:1.1rem; }
    .cert-tag   { background:#d9ead3; color:#2d6a2d; padding:2px 8px;
                  border-radius:4px; font-size:0.75rem; font-weight:600; }
    .warn-tag   { background:#fce5cd; color:#7f4c0a; padding:2px 8px;
                  border-radius:4px; font-size:0.75rem; font-weight:600; }
    .err-tag    { background:#fce4d6; color:#8b0000; padding:2px 8px;
                  border-radius:4px; font-size:0.75rem; font-weight:600; }
    .client-box { background:#e8f4e8; border:1px solid #2d6a2d; border-radius:8px;
                  padding:0.6rem 1rem; margin-bottom:0.8rem; font-weight:600; color:#2d6a2d; }
</style>
""", unsafe_allow_html=True)

# ── CARTELLA LISTINI ─────────────────────────────────────────────────────────
LISTINI_DIR = Path(__file__).parent / "listini"
LISTINI_DIR.mkdir(exist_ok=True)

CLIENTE_BUILTIN = "ALFA SRL ⭐"   # client con prezzi integrati


def listini_salvati():
    """Restituisce la lista di clienti con listino salvato su disco."""
    files = sorted(LISTINI_DIR.glob("*.xlsx"))
    return [f.stem for f in files]


def path_listino(nome_cliente: str) -> Path:
    return LISTINI_DIR / f"{nome_cliente}.xlsx"


# ── PREZZI ALFA INTEGRATI (fallback senza file) ──────────────────────────────
OPZIONI_ALFA = {
    'CGFG16R3G1,5B10':  [("1E.02.040.0045.a — Cavo FG16OR16 3x1,5mm² ✓", 2.76),
                          ("1E.02.040.0095.a — Cavo FG16R16 3x1,5mm² resist.fuoco", 3.14)],
    'CGFG16R3G2,5B10':  [("1E.02.040.0045.b — Cavo FG16OR16 3x2,5mm² ✓", 3.27),
                          ("1E.02.040.0095.b — Cavo FG16R16 3x2,5mm² resist.fuoco", 3.72)],
    'BEG250SE':         [("1E.06.070.0050 — App. LED emergenza IP65 ✓", 284.62)],
    'BEG40004H':        [("1E.06.020.0335.b — Plafoniera stagna 2x36W IP65", 89.84),
                          ("1E.06.020.0335.a — Plafoniera stagna 1x36W IP65", 82.84),
                          ("1E.06.020.0335.c — Plafoniera stagna 3x36W IP65", 96.84),
                          ("1E.06.060.0130.a — Plafoniera LED tonda est./int.", 124.07)],
    'ABBDS201LHC16AC30':[("1E.03.030.0300.a — Int.diff.magn. 1P+N 16A 30mA 4,5kA", 84.25),
                          ("1E.03.030.0300.b — Int.diff.magn. 1P+N 16A 30mA 3kA", 80.39)],
    'GEWGW66326N':      [("1E.05.010.0010.a — Presa CEE17 2P+T 16A (solo mat.)", 19.71),
                          ("1E.20.010.001 CAP — Presa CEE17 16A (mat.+posa)", 118.06),
                          ("1E.20.010.003 CAP — Presa CEE17 32A (mat.+posa)", 133.39)],
    'BDSGHZ15002':      [("1E.02.040.0230.a — Cavo mult. scherm. 300V 2 anime", 1.59),
                          ("1E.02.040.0085.a — Cavo bipol. scherm. 2x1,5", 2.55),
                          ("1E.02.040.0035.a — Cavo bipol. FG16 2x1,5mm²", 2.20)],
    'BDSAHZ75422M100':  [("1E.02.040.0230.c — Cavo mult. scherm. 4 anime", 1.85),
                          ("1E.02.040.0230.b — Cavo mult. scherm. 3 anime", 1.72),
                          ("1E.02.040.0230.a — Cavo mult. scherm. 2 anime", 1.59)],
    'BDSGHZ15024':      [("1E.02.042.0360.d — Cavo mult. scherm. RF 19-24 anime 1,5mm²", 15.59),
                          ("1E.02.042.0360.c — Cavo mult. scherm. RF 13-18 anime 1,5mm²", 12.42),
                          ("1E.02.042.0360.e — Cavo mult. scherm. RF 25-36 anime 1,5mm²", 18.05)],
    'SNRSL40120':       [("1E.02.030.0080.c — Canaletta PVC rigida (Ø20)", 10.12)],
    'SNRENN47930':      [("1E.01.040.0120 — Fascetta fissaggio (verificare)", 15.65)],
    'A1S1900.13/XFT':   [("1E.07.050.0170 — Anello passacavo metallico", 9.43)],
    'GEWGW44205':       [("1E.02.020.0010.e — Cassetta derivazione (verif. IP56)", 4.37)],
}

TARIFFE_MANO = [
    ("RU.00.01.00.0040 — Tecnico specialistico impiantista  €38,43/h", 38.43),
    ("RU.00.00.00.0000 — Operaio edile lv4° specializzato super  €41,68/h", 41.68),
    ("RU.00.00.00.0005 — Operaio edile lv3° specializzato  €39,50/h", 39.50),
    ("RU.00.00.00.0010 — Operaio edile lv2° qualificato  €36,71/h", 36.71),
    ("RU.00.00.00.0015 — Operaio edile lv1° comune  €32,98/h", 32.98),
    ("MA.00.060.0030   — Operaio impiantista 1° livello  €21,48/h", 21.48),
]

# ── PARSING LISTINO EXCEL ────────────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def parse_listino(listino_bytes: bytes):
    wb = openpyxl.load_workbook(io.BytesIO(listino_bytes), read_only=True, data_only=True)

    sheet_ep = None
    for name in wb.sheetnames:
        if "EP" in name.upper() or "COMPLETO" in name.upper() or "PREZZ" in name.upper():
            sheet_ep = wb[name]; break
    if sheet_ep is None:
        sheet_ep = wb[wb.sheetnames[0]]

    rows_ep = []
    header_found = False
    col_tar = col_ds = col_de = col_um = col_pr = None

    for row in sheet_ep.iter_rows(values_only=True):
        if not header_found:
            row_str = [str(c).upper().strip() if c else "" for c in row]
            for i, cell in enumerate(row_str):
                if "TARIFFA" in cell or "CODICE" in cell: col_tar = i
                if "SINTETICA" in cell or ("DESCR" in cell and col_ds is None): col_ds = i
                if "ESTESA" in cell: col_de = i
                if "U.M" in cell or "UM" in cell: col_um = i
                if "PREZZO" in cell or "PRICE" in cell: col_pr = i
            if col_tar is not None and col_pr is not None:
                header_found = True
            continue

        if col_tar is None or col_pr is None: continue
        tar = str(row[col_tar]).strip() if row[col_tar] else ""
        ds  = str(row[col_ds]).strip()  if col_ds is not None and row[col_ds] else ""
        de  = str(row[col_de]).strip()  if col_de is not None and row[col_de] else ""
        um  = str(row[col_um]).strip()  if col_um is not None and row[col_um] else ""
        try:
            pr = float(str(row[col_pr]).replace(",", ".").replace("€","").strip()) if row[col_pr] else 0.0
        except (ValueError, TypeError):
            pr = 0.0
        if tar and tar != "None" and pr > 0:
            descr = ds if ds and ds != "None" else de
            rows_ep.append({"tariffa": tar, "descr": descr, "um": um, "prezzo": pr})

    df_ep = pd.DataFrame(rows_ep)

    confermati = {}
    for fname in wb.sheetnames:
        if fname.upper() in ("FOGLIO1","FOGLIO3","SHEET1","SHEET3","CONFERMATI","CLIENTE"):
            sh = wb[fname]
            for row in sh.iter_rows(values_only=True):
                tar = str(row[0]).strip() if row[0] else ""
                ds  = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                try:
                    pr = float(str(row[4]).replace(",",".")) if len(row) > 4 and row[4] else 0.0
                except (ValueError, TypeError):
                    pr = 0.0
                if tar and tar != "None" and pr > 0:
                    confermati[tar] = (f"{tar} — {ds} ✓", pr)

    wb.close()
    return df_ep, confermati


def cerca_nel_listino(descrizione_articolo, df_ep, confermati, top_n=5):
    if df_ep.empty:
        return []
    stopwords = {"per","con","del","della","dei","degli","alle","alla","che","una",
                 "uno","gli","le","da","in","di","il","la","lo","e","a","non"}
    words = [w.lower() for w in re.split(r'[\s\-/,\.]+', descrizione_articolo)
             if len(w) >= 3 and not w.isdigit() and w.lower() not in stopwords]
    if not words:
        return []

    mask_1e = df_ep["tariffa"].str.startswith("1E", na=False)
    df_1e = df_ep[mask_1e].copy() if mask_1e.any() else df_ep.copy()

    descr_lower = df_1e["descr"].str.lower().fillna("")
    scores = sum(descr_lower.str.contains(re.escape(w), regex=True) for w in words)
    df_1e = df_1e.copy()
    df_1e["score"] = scores
    df_match = df_1e[df_1e["score"] > 0].sort_values("score", ascending=False).head(top_n)

    results = []
    seen = set()
    for _, row in df_match.iterrows():
        tar = row["tariffa"]
        if tar in confermati and tar not in seen:
            results.append(confermati[tar]); seen.add(tar)
    for _, row in df_match.iterrows():
        tar = row["tariffa"]
        if tar not in seen:
            results.append((f"{tar} — {row['descr']}", row["prezzo"])); seen.add(tar)
    return results


# ── PARSING PDF ─────────────────────────────────────────────────────────────
def parse_rapportini(pdf_bytes):
    manodopera, materiali = [], []
    EW_RE  = re.compile(r'^(EW\d+)\s+(.+?)\s+(ORE)\s+([\d,]+)$')
    MAT_RE = re.compile(r'^([A-Za-z0-9\.\,\/\*\+]+)\s+(.+?)\s+(PZ|M\b|KG|MT)\s+([\d,]+)$')
    DAY_RE = re.compile(r'Giorno\s+(\d{2}/\d{2}/\d{4})')
    def fl(s): return float(s.replace(',', '.'))
    SKIP = {"RAPPORTINO","GS IMPIANTI","Via Riso","20826","0296","Commessa n","Data:","Rif. cantiere",
            "IMPIANTI","Codice","ACCETTAZIONE","Firma","Cliente","ALFA","P.ZZA","21100","Cantiere","Tel.","P.IVA"}

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            lines = (page.extract_text() or "").splitlines()
            current_day = None; in_m = in_t = False; cew = None; dl = []
            for line in lines:
                line = line.strip()
                if not line or any(line.startswith(s) for s in SKIP): continue
                if "DISTINTA MANODOPERA" in line: in_m=True; in_t=False; continue
                if "DISTINTA MATERIALI"  in line: in_t=True; in_m=False; continue
                m = DAY_RE.search(line)
                if m: current_day = m.group(1); continue
                if not in_m and not in_t:
                    if line.startswith("EW"): in_m = True
                    elif re.match(r'^[A-Z][A-Z0-9]', line) and re.search(r'\s+(PZ|M\b)\s+[\d,]+', line): in_t = True
                if in_m:
                    m2 = EW_RE.match(line)
                    if m2:
                        if cew: manodopera.append({'data':current_day,'codice':cew['c'],'descrizione':" ".join(dl).strip(),'ore':cew['o']})
                        cew = {'c':m2.group(1),'o':fl(m2.group(4))}
                        d = m2.group(2).strip(); dl = [d] if d else []
                    elif cew and not line.startswith("ORDINE"): dl.append(line)
                if in_t and current_day:
                    m3 = MAT_RE.match(line)
                    if m3: materiali.append({'data':current_day,'codice':m3.group(1),'descrizione':m3.group(2).strip(),'um':m3.group(3),'qty':fl(m3.group(4))})
            if cew: manodopera.append({'data':current_day,'codice':cew['c'],'descrizione':" ".join(dl).strip(),'ore':cew['o']})

    mat_agg = {}
    for m in materiali:
        k = m['codice']
        if k not in mat_agg: mat_agg[k] = {'descrizione':m['descrizione'],'um':m['um'],'qty':0}
        mat_agg[k]['qty'] += m['qty']
    return manodopera, mat_agg


# ── GENERAZIONE EXCEL ────────────────────────────────────────────────────────
def genera_excel(cliente, commessa, periodo, manodopera, mat_agg, prezzi_mat,
                 voci_mat, tariffa_mano_label, tariffa_mano_val):
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "COMMESSA"
    ws.sheet_view.showGridLines = False

    BLU='1F3864'; LBLU='BDD7EE'; GVERDE='E2EFDA'; GR='F2F2F2'
    CERT='D9EAD3'; GIALL='FFF2CC'; ARN='FCE5CD'
    thin = Side(style='thin', color="AAAAAA")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hd(c, txt, dark=True, sz=9):
        c.value = txt
        c.font  = Font(bold=True, color="FFFFFF" if dark else "1F3864", size=sz, name="Calibri")
        c.fill  = PatternFill("solid", fgColor=BLU if dark else LBLU)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = brd

    def cs(c, val, bold=False, align="left", fmt=None, bg=None, sz=9, italic=False):
        c.value = val
        c.font  = Font(bold=bold, size=sz, name="Calibri", italic=italic)
        c.alignment = Alignment(horizontal=align, vertical="top", wrap_text=True)
        if fmt: c.number_format = fmt
        if bg:  c.fill = PatternFill("solid", fgColor=bg)
        c.border = brd

    for col, w in zip('ABCDEFG', [22,46,8,11,40,14,14]):
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:C4")
    ws["A1"].value = "GS IMPIANTI TECNOLOGICI SRL\nVia Risorgimento, 105/A — 20826 MISINTO (MB)\nTel. 0296328158 — P.IVA 07108690962"
    ws["A1"].font  = Font(bold=True, size=9, name="Calibri")
    ws["A1"].alignment = Alignment(vertical="top", wrap_text=True)
    ws.merge_cells("D1:G4")
    ws["D1"].value = f"SPETTABILE\n{cliente}\nP.IVA / Cod.Fisc. 03481930125"
    ws["D1"].font  = Font(bold=True, size=9, name="Calibri")
    ws["D1"].alignment = Alignment(vertical="top", wrap_text=True)
    ws.row_dimensions[1].height = 55

    ws.merge_cells("A5:G5")
    ws["A5"].value = f"  Commessa {commessa}   ·   {cliente} - MANUTENZIONE STRAORDINARIA IMPIANTI   ·   {periodo}"
    ws["A5"].font  = Font(bold=True, size=10, name="Calibri", color="FFFFFF")
    ws["A5"].fill  = PatternFill("solid", fgColor=BLU)
    ws["A5"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[5].height = 18; ws.row_dimensions[6].height = 6

    ws.merge_cells("A7:G7")
    ws["A7"].value = "  MATERIALI"
    ws["A7"].font  = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
    ws["A7"].fill  = PatternFill("solid", fgColor=BLU)
    ws["A7"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[7].height = 16

    r = 8
    for col, txt in enumerate(["Codice Articolo","Descrizione","U.M.","Quantità","Voce Listino","Prezzo €/u","Importo €"], 1):
        hd(ws.cell(r, col), txt, dark=False)
    ws.row_dimensions[r].height = 28

    mat_start = 9
    for i, (cod, v) in enumerate(mat_agg.items()):
        r = mat_start + i
        prezzo = prezzi_mat.get(cod)
        voce   = voci_mat.get(cod, "")
        is_cert = "✓" in voce
        if is_cert: bg = CERT
        elif prezzo is not None and prezzo > 0: bg = GIALL
        else: bg = 'FCE4D6'

        cs(ws.cell(r,1), cod, bg=bg, sz=8)
        cs(ws.cell(r,2), v['descrizione'], bg=bg, sz=8)
        cs(ws.cell(r,3), v['um'], align="center", bg=bg, sz=8)
        c=ws.cell(r,4); c.value=v['qty']; c.number_format='#,##0.00'
        c.font=Font(size=9,name="Calibri"); c.alignment=Alignment(horizontal="right"); c.border=brd
        if bg: c.fill=PatternFill("solid",fgColor=bg)
        cs(ws.cell(r,5), voce if voce else "DA INSERIRE MANUALMENTE", bg=bg, sz=8, italic=True)
        fc=ws.cell(r,6); fc.value=prezzo if prezzo else None; fc.number_format='#,##0.00'
        fc.alignment=Alignment(horizontal="right"); fc.font=Font(size=9,name="Calibri",bold=True)
        if bg: fc.fill=PatternFill("solid",fgColor=bg); fc.border=brd
        gc=ws.cell(r,7); gc.value=round(v['qty']*prezzo,2) if prezzo else None; gc.number_format='#,##0.00'
        gc.alignment=Alignment(horizontal="right"); gc.font=Font(size=9,name="Calibri",bold=True)
        if bg: gc.fill=PatternFill("solid",fgColor=bg); gc.border=brd
        ws.row_dimensions[r].height = 14

    mat_end = mat_start + len(mat_agg) - 1
    rt = mat_end + 1
    ws.merge_cells(f"A{rt}:F{rt}")
    ws[f"A{rt}"].value = "TOTALE MATERIALI"
    ws[f"A{rt}"].font  = Font(bold=True, size=9, name="Calibri")
    ws[f"A{rt}"].fill  = PatternFill("solid", fgColor=GVERDE)
    ws[f"A{rt}"].alignment = Alignment(horizontal="right", vertical="center")
    for col in range(1,7): ws.cell(rt,col).fill=PatternFill("solid",fgColor=GVERDE); ws.cell(rt,col).border=brd
    tot_mat = sum(v['qty']*prezzi_mat.get(cod,0) for cod,v in mat_agg.items() if prezzi_mat.get(cod))
    tc=ws.cell(rt,7); tc.value=round(tot_mat,2); tc.font=Font(bold=True,size=10,name="Calibri")
    tc.fill=PatternFill("solid",fgColor=GVERDE); tc.number_format='#,##0.00'
    tc.alignment=Alignment(horizontal="right"); tc.border=brd
    ws.row_dimensions[rt].height=18; ws.row_dimensions[rt+1].height=8

    mh = rt + 2
    ws.merge_cells(f"A{mh}:G{mh}")
    ws[f"A{mh}"].value = "  INTERVENTI / MANODOPERA"
    ws[f"A{mh}"].font  = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
    ws[f"A{mh}"].fill  = PatternFill("solid", fgColor=BLU)
    ws[f"A{mh}"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[mh].height = 16

    rh = mh + 1
    for col, txt in enumerate(["Data","Descrizione Intervento","Ore","","Categoria Tariffa Listino","Tariffa €/h","Importo €"], 1):
        hd(ws.cell(rh, col), txt, dark=False)
    ws.row_dimensions[rh].height = 28

    ms = rh + 1
    tot_ore = sum(m['ore'] for m in manodopera)
    for i, m in enumerate(manodopera):
        r = ms + i; bg = GR if i % 2 == 1 else None
        cs(ws.cell(r,1), m['data'], bg=bg, sz=8)
        cs(ws.cell(r,2), m['descrizione'], bg=bg, sz=8)
        c=ws.cell(r,3); c.value=m['ore']; c.number_format='#,##0.00'
        c.font=Font(size=9,name="Calibri"); c.alignment=Alignment(horizontal="right")
        if bg: c.fill=PatternFill("solid",fgColor=bg); c.border=brd
        cs(ws.cell(r,5), tariffa_mano_label, bg=ARN, sz=7, italic=True)
        fc=ws.cell(r,6); fc.value=tariffa_mano_val; fc.number_format='#,##0.00'
        fc.font=Font(size=9,name="Calibri"); fc.alignment=Alignment(horizontal="right")
        fc.fill=PatternFill("solid",fgColor=ARN); fc.border=brd
        gc=ws.cell(r,7); gc.value=round(m['ore']*tariffa_mano_val,2); gc.number_format='#,##0.00'
        gc.font=Font(size=9,name="Calibri",bold=True); gc.alignment=Alignment(horizontal="right")
        if bg: gc.fill=PatternFill("solid",fgColor=bg); gc.border=brd
        ws.row_dimensions[r].height = 26

    me = ms + len(manodopera) - 1
    rm = me + 1
    ws.merge_cells(f"A{rm}:F{rm}")
    ws[f"A{rm}"].value = f"TOTALE MANODOPERA  ({tot_ore:.0f} ore)"
    ws[f"A{rm}"].font  = Font(bold=True, size=9, name="Calibri")
    ws[f"A{rm}"].fill  = PatternFill("solid", fgColor=GVERDE)
    ws[f"A{rm}"].alignment = Alignment(horizontal="right", vertical="center")
    for col in range(1,7): ws.cell(rm,col).fill=PatternFill("solid",fgColor=GVERDE); ws.cell(rm,col).border=brd
    tot_mano = tot_ore * tariffa_mano_val
    tc2=ws.cell(rm,7); tc2.value=round(tot_mano,2); tc2.font=Font(bold=True,size=10,name="Calibri")
    tc2.fill=PatternFill("solid",fgColor=GVERDE); tc2.number_format='#,##0.00'
    tc2.alignment=Alignment(horizontal="right"); tc2.border=brd
    ws.row_dimensions[rm].height = 18

    rg = rm + 2; ws.row_dimensions[rg-1].height = 8
    ws.merge_cells(f"A{rg}:F{rg}")
    ws[f"A{rg}"].value = "TOTALE GENERALE"
    ws[f"A{rg}"].font  = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
    ws[f"A{rg}"].fill  = PatternFill("solid", fgColor=BLU)
    ws[f"A{rg}"].alignment = Alignment(horizontal="right", vertical="center")
    for col in range(1,7): ws.cell(rg,col).fill=PatternFill("solid",fgColor=BLU); ws.cell(rg,col).border=brd
    tg=ws.cell(rg,7); tg.value=round(tot_mat+tot_mano,2)
    tg.font=Font(bold=True,size=11,color="FFFFFF",name="Calibri")
    tg.fill=PatternFill("solid",fgColor=BLU); tg.number_format='#,##0.00'
    tg.alignment=Alignment(horizontal="right"); tg.border=brd
    ws.row_dimensions[rg].height = 24

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf.getvalue(), round(tot_mat+tot_mano, 2)


# ════════════════════════════════════════════════════════════════════════════
# SIDEBAR — SELEZIONE CLIENTE / GESTIONE LISTINI
# ════════════════════════════════════════════════════════════════════════════

with st.sidebar:
    st.markdown("## 🏢 Cliente")
    st.markdown("---")

    clienti_salvati = listini_salvati()
    opzioni = [CLIENTE_BUILTIN] + clienti_salvati + ["➕ Aggiungi nuovo cliente..."]

    scelta = st.selectbox(
        "Seleziona cliente",
        opzioni,
        label_visibility="collapsed",
        key="cliente_sel"
    )

    st.markdown("---")

    if scelta == "➕ Aggiungi nuovo cliente...":
        st.markdown("#### Nuovo cliente")
        nuovo_nome = st.text_input("Nome cliente (es. BETA SRL)", key="nuovo_nome")
        nuovo_file = st.file_uploader("Listino Excel (.xlsx)", type=["xlsx","xls"], key="nuovo_listino")
        if st.button("💾 Salva listino", type="primary", use_container_width=True):
            if not nuovo_nome.strip():
                st.error("Inserisci il nome del cliente.")
            elif not nuovo_file:
                st.error("Carica il file del listino.")
            else:
                dest = path_listino(nuovo_nome.strip())
                dest.write_bytes(nuovo_file.read())
                st.success(f"✅ Listino di **{nuovo_nome.strip()}** salvato!")
                st.rerun()
        st.stop()

    # Mostra il cliente attivo
    if scelta == CLIENTE_BUILTIN:
        st.markdown(f'<div class="client-box">✅ {CLIENTE_BUILTIN}<br><small>Prezzi integrati ALFA 2025</small></div>', unsafe_allow_html=True)
        df_ep, confermati = pd.DataFrame(), {}
        nome_cliente_display = "ALFA SRL"
        usa_builtin = True
    else:
        st.markdown(f'<div class="client-box">✅ {scelta}<br><small>Listino caricato</small></div>', unsafe_allow_html=True)
        p = path_listino(scelta)
        with st.spinner("Caricamento listino..."):
            df_ep, confermati = parse_listino(p.read_bytes())
        n_ep   = len(df_ep)
        n_conf = len(confermati)
        st.caption(f"📋 {n_ep:,} voci · {n_conf} già confermate")
        nome_cliente_display = scelta
        usa_builtin = False

        # Tasto per eliminare il listino
        with st.expander("⚙️ Gestione"):
            if st.button("🗑️ Elimina listino", type="secondary", use_container_width=True):
                path_listino(scelta).unlink(missing_ok=True)
                st.rerun()

    st.markdown("---")
    st.caption("GS Impianti Tecnologici\nGeneratore Commesse v2.0")


# ════════════════════════════════════════════════════════════════════════════
# AREA PRINCIPALE
# ════════════════════════════════════════════════════════════════════════════

st.markdown('<p class="main-title">⚡ GS Impianti Tecnologici</p>', unsafe_allow_html=True)
st.markdown('<p class="sub-title">Generatore automatico commesse mensili — da rapportini PDF a Excel cliente</p>', unsafe_allow_html=True)

# ── STEP 1: Upload PDF ───────────────────────────────────────────────────────
st.markdown('<div class="step-box"><span class="step-num">① Carica il PDF dei rapportini del mese</span></div>', unsafe_allow_html=True)

col1, col2, col3 = st.columns([2, 1, 1])
with col1:
    pdf_file = st.file_uploader("Rapportini PDF (da EDISON)", type=["pdf"], label_visibility="collapsed")
with col2:
    cliente  = st.text_input("Cliente", value=nome_cliente_display)
with col3:
    commessa = st.text_input("N. Commessa", value="26 0065")

periodo = st.text_input("Periodo", value="MARZO 2026", placeholder="es. APRILE 2026")

if not pdf_file:
    st.info("⬆️ Carica il PDF dei rapportini per iniziare")
    st.stop()

# ── PARSING ──────────────────────────────────────────────────────────────────
with st.spinner("Analisi rapportini in corso..."):
    manodopera, mat_agg = parse_rapportini(pdf_file.read())

if not manodopera and not mat_agg:
    st.error("Nessun dato trovato nel PDF. Verificare che sia il file corretto.")
    st.stop()

st.success(f"✅ Estratti **{len(manodopera)} righe manodopera** e **{len(mat_agg)} materiali** dai rapportini")
st.divider()

# ── STEP 2: MATERIALI ────────────────────────────────────────────────────────
st.markdown('<div class="step-box"><span class="step-num">② Verifica e scegli i prezzi dei materiali</span></div>', unsafe_allow_html=True)

hcol = st.columns([1.8, 3, 0.9, 0.7, 3.5, 1.2])
for hc, ht in zip(hcol, ["Codice","Descrizione","Qtà","UM","Voce listino","€/u"]):
    hc.markdown(f"<small><b>{ht}</b></small>", unsafe_allow_html=True)
st.markdown("<hr style='margin:2px 0; border-color:#1F3864'>", unsafe_allow_html=True)

prezzi_mat = {}
voci_mat   = {}
righe_da_verificare = []

for cod, v in mat_agg.items():
    if usa_builtin:
        opts = OPZIONI_ALFA.get(cod, [])
    else:
        opts = cerca_nel_listino(v['descrizione'], df_ep, confermati)

    with st.container():
        col_cod, col_desc, col_qty, col_um, col_scelta, col_prezzo = st.columns([1.8, 3, 0.9, 0.7, 3.5, 1.2])
        with col_cod:   st.markdown(f"**{cod}**")
        with col_desc:  st.markdown(f"<small>{v['descrizione']}</small>", unsafe_allow_html=True)
        with col_qty:   st.markdown(f"**{v['qty']:.0f}**")
        with col_um:    st.markdown(v['um'])

        if len(opts) == 0:
            with col_scelta:
                st.markdown('<span class="err-tag">⚠ Non nel listino</span>', unsafe_allow_html=True)
            with col_prezzo:
                p = st.number_input("€", min_value=0.0, value=0.0, step=0.01,
                                    key=f"p_{cod}", label_visibility="collapsed")
                prezzi_mat[cod] = p if p > 0 else None
                voci_mat[cod]   = ""
            righe_da_verificare.append(cod)

        elif len(opts) == 1:
            label, prezzo = opts[0]
            with col_scelta:
                tag = '<span class="cert-tag">✓ Confermato</span>' if "✓" in label else '<span class="warn-tag">◈ Da verificare</span>'
                st.markdown(f"{tag} <small>{label.replace(' ✓','')}</small>", unsafe_allow_html=True)
            with col_prezzo:
                p = st.number_input("€", min_value=0.0, value=float(prezzo), step=0.01,
                                    key=f"p_{cod}", label_visibility="collapsed")
                prezzi_mat[cod] = p
                voci_mat[cod]   = label
        else:
            with col_scelta:
                labels = [o[0] for o in opts]
                scelta_voce = st.selectbox("▼", labels, key=f"s_{cod}", label_visibility="collapsed")
                selected_p  = next(p for l,p in opts if l==scelta_voce)
                voci_mat[cod] = scelta_voce
            with col_prezzo:
                p = st.number_input("€", min_value=0.0, value=float(selected_p), step=0.01,
                                    key=f"p_{cod}", label_visibility="collapsed")
                prezzi_mat[cod] = p

    st.markdown("<hr style='margin:2px 0; border-color:#eee'>", unsafe_allow_html=True)

if righe_da_verificare:
    st.warning(f"⚠️ **{len(righe_da_verificare)} materiali** non trovati nel listino — inserire il prezzo manualmente.")

st.divider()

# ── STEP 3: MANODOPERA ───────────────────────────────────────────────────────
tot_ore_prev = sum(m['ore'] for m in manodopera)
st.markdown(f'<div class="step-box"><span class="step-num">③ Scegli la tariffa oraria per la manodopera</span><br><small>Si applica a tutte le {len(manodopera)} righe ({tot_ore_prev:.0f} ore totali)</small></div>', unsafe_allow_html=True)

col_t1, col_t2 = st.columns([4, 1])
with col_t1:
    tariffa_label = st.selectbox("Categoria tariffa", [t[0] for t in TARIFFE_MANO], label_visibility="collapsed")
tariffa_val = next(p for l,p in TARIFFE_MANO if l==tariffa_label)
with col_t2:
    tariffa_val = st.number_input("€/h", min_value=0.0, value=float(tariffa_val), step=0.50, label_visibility="collapsed")

tot_ore  = sum(m['ore'] for m in manodopera)
tot_mano = tot_ore * tariffa_val
tot_mat  = sum(v['qty'] * (prezzi_mat.get(cod) or 0) for cod,v in mat_agg.items())

st.markdown(f"""
<div style='background:#1F3864;color:white;padding:1rem 1.5rem;border-radius:8px;margin-top:1rem;'>
  <span style='font-size:1rem;'>
    📦 Materiali: <b>€ {tot_mat:,.2f}</b> &nbsp;&nbsp;
    👷 Manodopera: <b>€ {tot_mano:,.2f}</b> ({tot_ore:.0f} ore × €{tariffa_val:.2f}) &nbsp;&nbsp;
    💰 <span style='font-size:1.2rem;'>TOTALE: € {tot_mat+tot_mano:,.2f}</span>
  </span>
</div>
""", unsafe_allow_html=True)

st.divider()

# ── STEP 4: Genera Excel ─────────────────────────────────────────────────────
st.markdown('<div class="step-box"><span class="step-num">④ Genera il file Excel da inviare al cliente</span></div>', unsafe_allow_html=True)

if st.button("🗂️ Genera Excel Commessa", type="primary", use_container_width=True):
    with st.spinner("Generazione Excel..."):
        excel_bytes, totale = genera_excel(
            cliente, commessa, periodo,
            manodopera, mat_agg, prezzi_mat, voci_mat,
            tariffa_label, tariffa_val
        )
    nome_file = f"Commessa_{commessa.replace(' ','_')}_{periodo.replace(' ','_')}.xlsx"
    st.download_button(
        label=f"⬇️ Scarica {nome_file}  (Totale: € {totale:,.2f})",
        data=excel_bytes,
        file_name=nome_file,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
    st.balloons()
